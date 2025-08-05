import logging
from io import BytesIO
from typing import AsyncGenerator,List

import networkx as nx
import numpy as np
import openpyxl

from .base_parser import AsyncParser

logger = logging.getLogger(__name__)


class XLSXParser(AsyncParser[bytes]):
    """A parser for XLSX data that yields each row as a list of cell values."""

    def __init__(self):
        """Initializes the XLSXParser."""
        pass

    async def ingest(self, data: bytes, **kwargs) -> AsyncGenerator[List[str], None]:
        """Ingest XLSX data and yield a list of strings for each row."""
        if not isinstance(data, bytes):
            raise TypeError("XLSX data must be in bytes format.")

        try:
            workbook = openpyxl.load_workbook(filename=BytesIO(data))
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        yield [str(cell) if cell is not None else "" for cell in row]
        except Exception as e:
            logger.error(f"Failed to read or process XLSX file: {e}", exc_info=True)
            raise ValueError(f"Error processing XLSX file: {str(e)}") from e

class XLSXParserAdvanced(AsyncParser[bytes]):
    """
    An advanced XLSX parser that identifies and extracts distinct data tables
    (connected components of non-empty cells) from each sheet.
    """

    def __init__(self):
        """Initializes the XLSXParserAdvanced."""
        pass

    def _find_connected_components(self, arr: np.ndarray):
        """Identifies and yields connected components from a numpy array of cells."""
        graph = nx.grid_2d_graph(arr.shape[0], arr.shape[1])
        
        # Remove nodes corresponding to empty cells
        empty_cells = list(zip(*np.where(arr == None), strict=False))
        graph.remove_nodes_from(empty_cells)
        
        for component in nx.connected_components(graph):
            if not component:
                continue
            rows, cols = zip(*component, strict=False)
            min_row, max_row = min(rows), max(rows)
            min_col, max_col = min(cols), max(cols)
            yield arr[min_row : max_row + 1, min_col : max_col + 1].astype(str)

    async def ingest(self, data: bytes, **kwargs) -> AsyncGenerator[str, None]:
        """Ingest XLSX data and yield text from each identified data table."""
        if not isinstance(data, bytes):
            raise TypeError("XLSX data must be in bytes format.")

        try:
            workbook = openpyxl.load_workbook(filename=BytesIO(data))
            for ws in workbook.worksheets:
                # Convert sheet to numpy array for processing
                ws_data = np.array([[cell.value for cell in row] for row in ws.iter_rows()], dtype=object)
                
                for table in self._find_connected_components(ws_data):
                    if table.size == 0:
                        continue
                    
                    # Yield each row of the identified table as a string
                    for row in table:
                        yield ", ".join(row)
        except Exception as e:
            logger.error(f"Failed to read or process XLSX file with advanced parser: {e}", exc_info=True)
            raise ValueError(f"Error processing XLSX file with advanced parser: {str(e)}") from e